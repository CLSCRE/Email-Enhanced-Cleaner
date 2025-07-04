
import streamlit as st
import pandas as pd
import requests
import time
from io import BytesIO
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

API_KEY = st.secrets["EMAILABLE_API_KEY"]
BASE_URL = "https://api.emailable.com/v1/verify"

def extract_emails(df):
    email_cols = [col for col in df.columns if 'email' in col.lower()]
    raw_emails = df[email_cols].values.flatten()
    emails = pd.Series(raw_emails).dropna().astype(str).str.strip().str.lower()
    return emails.drop_duplicates(), email_cols

def evaluate_score(score):
    try:
        score = int(score)
        if score >= 90:
            return ("Very Low", "Send")
        elif score >= 70:
            return ("Low", "Safe to Send")
        elif score >= 50:
            return ("Medium", "Review")
        elif score >= 30:
            return ("High", "Do Not Send")
        else:
            return ("Very High", "Do Not Send")
    except:
        return ("Unknown", "Review")

def classify_status(deliverable):
    if deliverable is True:
        return "Valid"
    elif deliverable is False:
        return "Invalid"
    else:
        return "Unknown"

def enrich_email(email):
    try:
        response = requests.get(BASE_URL, params={
            'api_key': API_KEY,
            'email': email
        })
        data = response.json()
        score = data.get('score')
        risk_level, action = evaluate_score(score)
        deliverable = data.get('deliverable')
        status = classify_status(deliverable)
        return {
            'Email': email,
            'Valid Format': data.get('format'),
            'Deliverable': deliverable,
            'MX Found': data.get('mx'),
            'SMTP Check': data.get('smtp'),
            'Is Free Email': data.get('free'),
            'Reason': data.get('reason'),
            'Status': status,
            'State': 'Deliverable' if status == 'Valid' else 'Undeliverable' if status == 'Invalid' else 'Risky'
        }
    except Exception as e:
        return {'Email': email, 'Error': str(e), 'Status': 'Error'}

st.set_page_config(page_title="CLS CRE Email Enrichment Tool", layout="wide")
st.image("https://clscre.com/wp-content/uploads/2023/05/CLS-CRE_logo_white.png", width=200)
st.markdown("<h4>📧 Email Address Enrichment Tool</h4>", unsafe_allow_html=True)
st.caption("Upload a spreadsheet with email addresses. We'll flag risky or invalid ones and show you which to fix.")

uploaded_file = st.file_uploader("Upload Excel or CSV File", type=["xlsx", "xls", "csv"])

if uploaded_file:
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    emails, email_cols = extract_emails(df)
    st.success(f"Found {len(emails)} unique email addresses.")

    enriched = []
    progress = st.progress(0)
    for i, email in enumerate(emails):
        enriched.append(enrich_email(email))
        progress.progress((i + 1) / len(emails))
        time.sleep(1)

    enriched_df = pd.DataFrame(enriched)
    st.dataframe(enriched_df)

    risky_emails = enriched_df[enriched_df['Status'].isin(['Invalid', 'Unknown', 'Error'])]['Email'].tolist()
    reason_map = enriched_df.set_index("Email")["Reason"].to_dict()
    styled_df = df.copy()

    output = BytesIO()
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    # First tab: Original Highlights
    ws_original = wb.create_sheet("Original Highlights")
    for r_idx, row in enumerate(dataframe_to_rows(styled_df, index=False, header=True), 1):
        ws_original.append(row)
        if r_idx == 1:
            continue
        for col in email_cols:
            email = styled_df.iloc[r_idx - 2][col]
            email_lower = str(email).strip().lower()
            cell = ws_original.cell(row=r_idx, column=styled_df.columns.get_loc(col) + 1)
            if email_lower in risky_emails:
                cell.font = Font(color="FF0000")
            elif reason_map.get(email_lower, "").lower() == "accepted_email":
                cell.font = Font(color="00AA00")

    # Second tab: Enriched Emails
    ws_enriched = wb.create_sheet("Enriched Emails")
    for r in dataframe_to_rows(enriched_df, index=False, header=True):
        ws_enriched.append(r)

    wb.save(output)
    st.download_button("📥 Download Results Excel", output.getvalue(), file_name="email_enrichment_final_output.xlsx")
